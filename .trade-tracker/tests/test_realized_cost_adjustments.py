from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch


TOOLS_DIR = Path(__file__).resolve().parents[1] / "tools"
sys.path.insert(0, str(TOOLS_DIR))

from install_preview_service import RUNTIME_PACKAGES
from trade_tracker import market_data as market_data_module
from trade_tracker.market_data import (
    eastmoney_quote_from_row,
    fetch_option_quote,
    fetch_tencent_fx_rates_to_cny,
    fetch_yahoo_fx_rates_to_cny,
    parse_hkex_option_detail,
    tencent_quote_from_payload,
    yahoo_quote_from_result,
)
from trade_tracker.html_tables import add_balanced_summary_table_script, normalize_legacy_holdings_table, normalize_legacy_open_option_sections
from trade_tracker.dividends import DIVIDEND_SHEET_NAME, load_dividend_events, load_workbook_dividend_events
from trade_tracker.options import build_stock_realized_income_maps, open_option_mark_for_row, patch_dashboard_data_with_options
from trade_tracker import state


class Cell:
    def __init__(self, raw):
        self.raw = raw


def row(**values):
    columns = {
        "kind": 1,
        "open_date": 2,
        "exp": 3,
        "close_date": 4,
        "ticker": 5,
        "event": 6,
        "strike": 7,
        "qty": 8,
        "open_price": 9,
        "close_price": 10,
        "fee": 11,
        "capital": 12,
        "pnl": 13,
        "multiplier": 19,
        "currency": 20,
    }
    return {columns[key]: Cell(value) for key, value in values.items()}


class FakeCore:
    @staticmethod
    def raw_text(value):
        return "" if value is None else str(value)

    @staticmethod
    def normalize_ticker(ticker, currency=""):
        return str(ticker or "").strip().upper()

    @staticmethod
    def normalize_currency(currency):
        mapping = {"人民币": "CNY", "港币": "HKD", "美元": "USD"}
        return mapping.get(str(currency or "").strip(), str(currency or "").strip().upper())

    @staticmethod
    def infer_currency_from_ticker(ticker):
        return "CNY"

    @staticmethod
    def lookup_security_name(ticker, currency="", allow_online=False):
        return str(ticker or "").strip().upper()

    @staticmethod
    def compute_row_metrics(cells):
        return {"pnl": float(cells[13].raw)}

    @staticmethod
    def row_capital(cells, trade_type, event):
        return float(cells.get(12, Cell(0)).raw or 0)


class RealizedCostAdjustmentTests(unittest.TestCase):
    def test_closed_stock_rows_are_grouped_by_ticker_and_currency(self):
        rows = [
            (
                2,
                row(
                    kind="股票",
                    open_date=46000,
                    close_date=46005,
                    ticker="TICKER_A",
                    event="现股",
                    qty=200,
                    open_price=10,
                    close_price=12,
                    fee=5,
                    pnl=395,
                    currency="人民币",
                ),
            ),
            (
                3,
                row(
                    kind="股票",
                    open_date=46006,
                    ticker="TICKER_A",
                    event="现股",
                    qty=100,
                    open_price=11,
                    pnl=999,
                    currency="人民币",
                ),
            ),
            (
                4,
                row(
                    kind="卖出",
                    open_date=46007,
                    close_date=46008,
                    ticker="TICKER_A",
                    event="认购",
                    qty=1,
                    open_price=0.2,
                    close_price=0,
                    fee=3,
                    pnl=999,
                    multiplier=100,
                    currency="人民币",
                ),
            ),
        ]

        self.assertEqual(build_stock_realized_income_maps(FakeCore(), rows), {("TICKER_A", "人民币"): 395.0})

    def test_patch_dashboard_data_reduces_long_holding_cost_and_syncs_summaries(self):
        current_year = str(date.today().year)
        prior_year = str(date.today().year - 1)
        closed_stock = row(
            kind="股票",
            open_date=46000,
            close_date=46005,
            ticker="TICKER_A",
            event="现股",
            qty=200,
            open_price=10,
            close_price=12,
            fee=5,
            pnl=395,
            currency="人民币",
        )
        data = {
            "holdings": [
                {
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "side": "多头",
                    "qty": "1000",
                    "all_in_cost": "人民币 10,000.00",
                    "avg_cost": "人民币 10.00",
                    "market_value": "人民币 11,500.00",
                    "last_price": "人民币 11.50",
                    "float_pnl": "人民币 1,500.00",
                    "daily_pnl": "人民币 120.00",
                }
            ],
            "stock_summary": [
                {
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "dividend": "人民币 0.00",
                    "unrealized_pnl": "人民币 1,500.00",
                    "total_pnl": "人民币 1,895.00",
                    "capital_raw": 10_000.0,
                    "capital_days_raw": 100_000.0,
                }
            ],
            "annual_summary": [
                {
                    "year": current_year,
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "dividend": "人民币 0.00",
                    "unrealized_pnl": "人民币 1,500.00",
                    "total_pnl": "人民币 1,895.00",
                    "capital_raw": 10_000.0,
                    "capital_days_raw": 100_000.0,
                },
                {
                    "year": prior_year,
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "dividend": "人民币 0.00",
                    "unrealized_pnl": "人民币 -50.00",
                    "total_pnl": "人民币 -50.00",
                    "capital_raw": 1_000.0,
                    "capital_days_raw": 10_000.0,
                },
            ],
        }

        patched = patch_dashboard_data_with_options(FakeCore(), [(2, closed_stock)], data)
        holding = patched["holdings"][0]

        self.assertEqual(holding["all_in_cost"], "人民币 9,605.00")
        self.assertEqual(holding["avg_cost"], "人民币 9.61")
        self.assertEqual(holding["float_pnl"], "人民币 1,895.00")
        self.assertEqual(holding["breakeven"], "-")
        self.assertEqual(patched["cost_text"], "人民币 9,605.00")
        self.assertEqual(patched["unrealized_pnl_text"], "人民币 1,895.00")
        self.assertEqual(patched["stock_summary"][0]["total_pnl"], "人民币 1,895.00")
        self.assertEqual(patched["annual_summary"][0]["total_pnl"], "人民币 1,895.00")
        self.assertEqual(patched["annual_summary"][1]["total_pnl"], "人民币 -50.00")

    def test_dividend_income_reduces_current_holding_cost_without_double_counting(self):
        current_year = str(date.today().year)
        data = {
            "holdings": [
                {
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "side": "多头",
                    "qty": "1000",
                    "all_in_cost": "人民币 10,000.00",
                    "avg_cost": "人民币 10.00",
                    "market_value": "人民币 10,500.00",
                    "last_price": "人民币 10.50",
                    "float_pnl": "人民币 500.00",
                    "daily_pnl": "人民币 0.00",
                }
            ],
            "stock_summary": [
                {
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "dividend": "人民币 200.00",
                    "unrealized_pnl": "人民币 500.00",
                    "total_pnl": "人民币 700.00",
                    "capital_raw": 10_000.0,
                    "capital_days_raw": 100_000.0,
                }
            ],
            "annual_summary": [
                {
                    "year": current_year,
                    "ticker": "TICKER_A",
                    "currency": "人民币",
                    "dividend": "人民币 200.00",
                    "unrealized_pnl": "人民币 500.00",
                    "total_pnl": "人民币 700.00",
                    "capital_raw": 10_000.0,
                    "capital_days_raw": 100_000.0,
                }
            ],
        }

        patched = patch_dashboard_data_with_options(FakeCore(), [], data)
        holding = patched["holdings"][0]

        self.assertEqual(holding["all_in_cost"], "人民币 9,800.00")
        self.assertEqual(holding["avg_cost"], "人民币 9.80")
        self.assertEqual(holding["float_pnl"], "人民币 700.00")
        self.assertEqual(holding["breakeven"], "-")
        self.assertEqual(patched["stock_summary"][0]["dividend"], "人民币 200.00")
        self.assertEqual(patched["stock_summary"][0]["total_pnl"], "人民币 700.00")
        self.assertEqual(patched["annual_summary"][0]["total_pnl"], "人民币 700.00")

    def test_workbook_dividend_sheet_is_loaded_as_events(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "tracker.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = DIVIDEND_SHEET_NAME
            sheet.append(["日期", "代码", "名称", "事件", "金额", "币种", "来源", "原始行"])
            sheet.append(["2026-02-03", "600000", "示例银行", "除权除息", 123.45, "人民币", "示例券商.xlsx", 8])
            workbook.save(path)

            events = load_workbook_dividend_events(FakeCore(), path)

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["ticker"], "600000")
        self.assertEqual(events[0]["currency"], "CNY")
        self.assertEqual(events[0]["amount"], 123.45)
        self.assertEqual(events[0]["kind"], "除权除息")

    def test_workbook_dividend_sheet_overrides_history_imports(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "tracker.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = DIVIDEND_SHEET_NAME
            sheet.append(["日期", "代码", "名称", "事件", "金额", "币种", "来源", "原始行"])
            sheet.append(["2026-02-03", "600000", "示例银行", "除权除息", 123.45, "人民币", "手工分红", 8])
            workbook.save(path)

            with patch("trade_tracker.dividends.load_history_dividend_events", return_value=[
                {"ticker": "000617", "name": "中油资本", "currency": "CNY", "amount": 61.56, "serial": 46000, "kind": "除权除息"}
            ]) as history:
                events = load_dividend_events(FakeCore(), path)

        history.assert_not_called()
        self.assertEqual([event["ticker"] for event in events], ["600000"])

    def test_short_holding_uses_inverse_cost_direction(self):
        closed_short = row(
            kind="卖出",
            open_date=46000,
            close_date=46003,
            ticker="TICKER_B",
            event="现股",
            qty=100,
            open_price=10,
            close_price=9,
            fee=2,
            pnl=98,
            currency="人民币",
        )
        data = {
            "holdings": [
                {
                    "ticker": "TICKER_B",
                    "currency": "人民币",
                    "side": "空头",
                    "qty": "-100",
                    "all_in_cost": "人民币 1,000.00",
                    "avg_cost": "人民币 10.00",
                    "market_value": "人民币 -850.00",
                    "last_price": "人民币 8.50",
                    "float_pnl": "人民币 150.00",
                    "daily_pnl": "人民币 0.00",
                }
            ],
            "stock_summary": [],
            "annual_summary": [],
        }

        patched = patch_dashboard_data_with_options(FakeCore(), [(2, closed_short)], data)
        holding = patched["holdings"][0]

        self.assertEqual(holding["all_in_cost"], "人民币 1,098.00")
        self.assertEqual(holding["avg_cost"], "人民币 10.98")
        self.assertEqual(holding["float_pnl"], "人民币 248.00")
        self.assertEqual(holding["breakeven"], "-")

    def test_losing_short_holding_shows_downside_breakeven_space(self):
        closed_short_loss = row(
            kind="卖出",
            open_date=46000,
            close_date=46003,
            ticker="TICKER_B",
            event="现股",
            qty=100,
            open_price=10,
            close_price=11,
            fee=2,
            pnl=-50,
            currency="人民币",
        )
        data = {
            "holdings": [
                {
                    "ticker": "TICKER_B",
                    "currency": "人民币",
                    "side": "空头",
                    "qty": "-100",
                    "all_in_cost": "人民币 1,000.00",
                    "avg_cost": "人民币 10.00",
                    "market_value": "人民币 -1,200.00",
                    "last_price": "人民币 12.00",
                    "float_pnl": "人民币 -200.00",
                    "daily_pnl": "人民币 0.00",
                }
            ],
            "stock_summary": [],
            "annual_summary": [],
        }

        patched = patch_dashboard_data_with_options(FakeCore(), [(2, closed_short_loss)], data)
        holding = patched["holdings"][0]

        self.assertEqual(holding["float_pnl"], "人民币 -250.00")
        self.assertEqual(holding["breakeven"], "-20.83%")

    def test_open_cash_secured_put_mark_uses_option_quote(self):
        open_put = row(
            kind="卖出",
            open_date=46140,
            exp=46170,
            ticker="TICKER_A",
            event="认沽",
            strike=10,
            qty=1,
            open_price=0.5,
            fee=2,
            capital=1000,
            multiplier=100,
            currency="人民币",
        )

        with patch("trade_tracker.options.fetch_option_quote", return_value={"option_code": "OPT", "last_price": 0.2}):
            _key, mark = open_option_mark_for_row(FakeCore(), open_put, {})

        self.assertEqual(mark["current_price"], "0.2")
        self.assertEqual(mark["float_pnl"], "28.00")
        self.assertEqual(mark["float_pnl_class"], "value-positive")
        self.assertEqual(mark["capital"], "1,000.00")

    def test_parse_hkex_option_detail_price(self):
        html = """
        <span class="floatleft col1a"><strong>Last Traded Price<br />
        (As of 13:30:00)</strong></span>
        <span class="floatright col1b"><strong>0.090</strong></span>
        """
        price, as_of = parse_hkex_option_detail(html)
        self.assertEqual(price, 0.09)
        self.assertEqual(as_of, "13:30:00")

    def test_parse_hkex_option_detail_prefers_bid_ask_midpoint(self):
        html = """
        Last Traded Price (As of 14:31:00) 0.370
        Bid / Ask 0.360 / 0.380
        """
        price, as_of = parse_hkex_option_detail(html)
        self.assertAlmostEqual(price, 0.37)
        self.assertEqual(as_of, "14:31:00")

    def test_public_quote_parsers_support_batch_sources(self):
        eastmoney_quote = eastmoney_quote_from_row(
            FakeCore(),
            ("00700", "HKD"),
            {"f14": "腾讯控股", "f2": 468800, "f1": 3, "f18": 479200},
        )
        self.assertEqual(eastmoney_quote["name"], "腾讯控股")
        self.assertEqual(eastmoney_quote["last_price"], 468.8)
        self.assertEqual(eastmoney_quote["prev_close"], 479.2)

        tencent_quote = tencent_quote_from_payload(
            FakeCore(),
            ("TSLA", "USD"),
            "200~特斯拉~TSLA.OQ~372.80~376.02",
        )
        self.assertEqual(tencent_quote["name"], "特斯拉")
        self.assertEqual(tencent_quote["last_price"], 372.8)
        self.assertEqual(tencent_quote["prev_close"], 376.02)

        yahoo_quote = yahoo_quote_from_result(
            FakeCore(),
            ("AAPL", "USD"),
            {"symbol": "AAPL", "shortName": "Apple Inc.", "regularMarketPrice": 269.53, "regularMarketPreviousClose": 270.17},
        )
        self.assertEqual(yahoo_quote["name"], "Apple Inc.")
        self.assertEqual(yahoo_quote["last_price"], 269.53)
        self.assertEqual(yahoo_quote["prev_close"], 270.17)

    def test_runtime_and_option_quotes_use_public_sources_only(self):
        self.assertEqual(RUNTIME_PACKAGES, ["openpyxl==3.1.5", "pandas==3.0.2"])
        with patch("trade_tracker.market_data.fetch_hkex_option_quote", return_value=None):
            self.assertIsNone(fetch_option_quote(FakeCore(), "DEMO", "港币", "认购", "2026-05-28", 32))

    def test_yahoo_fx_fallback_fetches_missing_rates_together(self):
        def fake_yahoo_rate(symbol):
            return {"HKDCNY=X": 0.92, "USDCNY=X": 7.20}.get(symbol)

        with patch("trade_tracker.market_data.fetch_yahoo_fx_rate_to_cny", side_effect=fake_yahoo_rate):
            rates = fetch_yahoo_fx_rates_to_cny({"港币": "HKDCNY=X", "美元": "USDCNY=X", "空": ""})

        self.assertEqual(rates, {"港币": 0.92, "美元": 7.20})

    def test_tencent_fx_rates_parse_batch_payload(self):
        payload = (
            'v_fxHKDCNY="310~港币人民币~HKDCNY~0.8716~0~20260501013139";\n'
            'v_fxUSDCNY="310~美元人民币~USDCNY~6.8299~0~20260501013144";\n'
        ).encode("gb18030")

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_):
                return False

            def read(self):
                return payload

        with patch("trade_tracker.market_data.urlopen", return_value=FakeResponse()):
            rates = fetch_tencent_fx_rates_to_cny({"港币": "fxHKDCNY", "美元": "fxUSDCNY"})

        self.assertEqual(rates, {"港币": 0.8716, "美元": 6.8299})

    def test_fx_rates_return_after_complete_tencent_rates(self):
        with (
            patch("trade_tracker.market_data.fetch_tencent_fx_rates_to_cny", return_value={"港币": 0.90, "美元": 7.10}) as tencent,
            patch("trade_tracker.market_data.fetch_eastmoney_fx_rates_to_cny", return_value={"港币": 0.91}) as eastmoney,
            patch("trade_tracker.market_data.fetch_yahoo_fx_rates_to_cny", return_value={"港币": 0.92, "美元": 7.20}) as yahoo,
        ):
            rates = market_data_module.compute_fx_rates_to_cny()

        self.assertEqual(rates["港币"], 0.90)
        self.assertEqual(rates["美元"], 7.10)
        tencent.assert_called_once()
        eastmoney.assert_not_called()
        yahoo.assert_not_called()

    def test_fx_rates_fallback_only_fetches_missing_tencent_rates(self):
        with (
            patch("trade_tracker.market_data.fetch_tencent_fx_rates_to_cny", return_value={"港币": 0.90}) as tencent,
            patch("trade_tracker.market_data.fetch_eastmoney_fx_rates_to_cny", return_value={}) as eastmoney,
            patch("trade_tracker.market_data.fetch_yahoo_fx_rates_to_cny", return_value={"美元": 7.20}) as yahoo,
        ):
            rates = market_data_module.compute_fx_rates_to_cny()

        self.assertEqual(rates["港币"], 0.90)
        self.assertEqual(rates["美元"], 7.20)
        tencent.assert_called_once()
        eastmoney.assert_called_once_with({"美元": "133.USDCNH"})
        yahoo.assert_called_once_with({"美元": "USDCNY=X"})

    def test_fx_prefetch_result_is_reused(self):
        previous_rates = market_data_module._FX_RATES_TO_CNY
        previous_future = market_data_module._FX_RATES_FUTURE
        market_data_module._FX_RATES_TO_CNY = None
        market_data_module._FX_RATES_FUTURE = None
        try:
            with patch("trade_tracker.market_data.compute_fx_rates_to_cny", return_value={"人民币": 1.0, "美元": 7.2}) as compute:
                market_data_module.start_fx_rates_prefetch()
                self.assertEqual(market_data_module.current_fx_rates_to_cny()["美元"], 7.2)
                self.assertEqual(market_data_module.current_fx_rates_to_cny()["美元"], 7.2)
                compute.assert_called_once()
        finally:
            market_data_module._FX_RATES_TO_CNY = previous_rates
            market_data_module._FX_RATES_FUTURE = previous_future

    def test_legacy_public_holdings_table_is_normalized_to_full_columns(self):
        html = """
        <table class="summary-table">
        <thead><tr><th>代码</th><th>数量</th><th>最新市值</th><th>持仓成本</th><th>浮动盈亏</th><th>现价</th><th>开仓</th></tr></thead>
        <tbody><tr><td>DEMO</td><td>3</td><td>美元 1,985.13</td><td>美元 1,986.00</td><td>美元 -0.87</td><td>661.71</td><td>2026-04-30</td></tr></tbody>
        </table>
        """
        normalized = normalize_legacy_holdings_table(FakeCore(), html)

        for header in ("名称", "盈亏率", "持股数", "当日盈亏", "个股仓位", "持股天数", "持仓均价", "回本空间", "方向", "币种", "最近买入"):
            self.assertIn(f">{header}</th>", normalized)
        self.assertIn(">662<", normalized)
        self.assertIn(">-0.04%</td>", normalized)
        self.assertIn(">+0.04%</td>", normalized)
        self.assertIn("value-positive", normalized)
        self.assertIn(">多头</td>", normalized)
        self.assertIn(">美元</td>", normalized)

    def test_legacy_public_open_option_table_is_normalized_to_full_columns(self):
        html = """
        <section>
        <h2 class="section-title">未平仓期权</h2>
        <table class="summary-table">
        <thead><tr><th>代码</th><th>事件</th><th>到期</th><th>行权价</th><th>数量</th><th>开仓价</th><th>现价</th><th>未实现盈亏</th></tr></thead>
        <tbody><tr><td>DEMO</td><td>认沽</td><td>2026-05-29</td><td>80</td><td>-3</td><td>6.413</td><td>9.82</td><td>美元 -1,021.95</td></tr></tbody>
        </table>
        </section>
        """
        normalized = normalize_legacy_open_option_sections(FakeCore(), html)

        for header in ("名称", "类型", "到期日", "乘数", "浮动盈亏", "占用本金", "币种"):
            self.assertIn(f">{header}</th>", normalized)
        self.assertIn(">卖出</td>", normalized)
        self.assertIn(">3</td>", normalized)
        self.assertIn(">100</td>", normalized)
        self.assertIn(">24,000.00</td>", normalized)
        self.assertIn(">美元</td>", normalized)

    def test_normalized_option_table_preserves_prefetched_marks(self):
        html = """
        <section>
        <h2 class="section-title">未平仓期权</h2>
        <p>只要期权那一行还没有平仓价，就会继续留在这里，方便你盯到期日。</p>
        <table class="summary-table">
        <thead><tr><th>代码</th><th>名称</th><th>类型</th><th>事件</th><th>到期日</th><th>行权价</th><th>数量</th><th>乘数</th><th>开仓价</th><th>币种</th></tr></thead>
        <tbody><tr><td>DEMO</td><td>示例标的</td><td>卖出</td><td>认购</td><td>2026/05/28</td><td>32</td><td>5</td><td>1000</td><td>0.55</td><td>港币</td></tr></tbody>
        </table>
        </section>
        """
        previous = dict(state.OPEN_OPTION_MARKS)
        state.OPEN_OPTION_MARKS = {
            ("DEMO", "卖出", "认购", "2026/05/28", "32", "5", "1000", "0.55", "港币"): {
                "current_price": "0.38",
                "float_pnl": "814.50",
                "float_pnl_class": "value-positive",
                "capital": "196,060.00",
            }
        }
        try:
            normalized = normalize_legacy_open_option_sections(FakeCore(), html)
        finally:
            state.OPEN_OPTION_MARKS = previous

        self.assertIn(">0.38</td>", normalized)
        self.assertIn(">814.50</td>", normalized)
        self.assertIn(">196,060.00</td>", normalized)
        self.assertIn("HKEX 等公开行情源", normalized)
        self.assertNotIn("本地行情客户端", normalized)

    def test_summary_table_tone_script_covers_legacy_float_profit_labels(self):
        html = """
        <html><body>
        <table class="summary-table">
        <thead><tr><th>代码</th><th>浮盈</th><th>成本</th></tr></thead>
        <tbody><tr><td>DEMO</td><td>123.45</td><td>1,000.00</td></tr></tbody>
        </table>
        </body></html>
        """
        patched = add_balanced_summary_table_script(html)

        self.assertIn("function applySummaryTableTones", patched)
        self.assertIn("浮盈|盈亏|收益|分红|年化|回本空间", patched)
        self.assertNotIn("normalized === '回本空间'", patched)


if __name__ == "__main__":
    unittest.main()
