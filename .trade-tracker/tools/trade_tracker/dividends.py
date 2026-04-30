from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from .names import iter_history_workbooks
from .utils import clean_name, clean_text, parse_float


DIVIDEND_SHEET_NAME = "分红记录"
DIVIDEND_EVENTS = {"除权除息", "股息个税征收", "分红", "派息", "股息", "Dividend", "dividend"}
BASE_DATE = date(1899, 12, 30)


def _load_workbook():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    return load_workbook


def _date_from_value(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for separator in ("-", "/", "."):
        parts = text.split(separator)
        if len(parts) != 3:
            continue
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except ValueError:
            continue
    return None


def _serial_from_date(value) -> float:
    parsed = _date_from_value(value)
    if parsed:
        return float((parsed - BASE_DATE).days)
    numeric = parse_float(value)
    return float(numeric or 0.0)


def _format_date(serial: float) -> str:
    if not serial:
        return ""
    try:
        return (BASE_DATE + timedelta(days=int(serial))).strftime("%Y/%m/%d")
    except (OverflowError, ValueError):
        return ""


def _normalize_currency(core, ticker: str, currency: object = "") -> str:
    normalized = clean_text(core.normalize_currency(currency))
    if normalized in {"CNY", "HKD", "USD"}:
        return normalized
    try:
        inferred = clean_text(core.infer_currency_from_ticker(ticker))
    except Exception:
        inferred = ""
    return inferred or normalized or "CNY"


def _source_key(event: dict[str, object]) -> tuple[object, ...]:
    source = clean_text(event.get("_source"))
    source_row = clean_text(event.get("_source_row"))
    if source or source_row:
        return (source, source_row)
    return (
        clean_text(event.get("ticker")),
        clean_text(event.get("currency")),
        round(float(event.get("serial") or 0.0), 6),
        clean_text(event.get("kind")),
        round(float(event.get("amount") or 0.0), 6),
    )


def _public_event(event: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in event.items() if not str(key).startswith("_")}


def _event(core, *, date_value, ticker, name, kind, amount, currency="", source="", source_row="") -> dict[str, object] | None:
    normalized_ticker = clean_text(core.normalize_ticker(ticker, currency))
    value = parse_float(amount)
    if not normalized_ticker or value is None or abs(value) < 1e-9:
        return None
    serial = _serial_from_date(date_value)
    normalized_currency = _normalize_currency(core, normalized_ticker, currency)
    normalized_name = clean_name(name)
    if not normalized_name:
        try:
            normalized_name = clean_name(core.lookup_security_name(normalized_ticker, normalized_currency))
        except Exception:
            normalized_name = ""
    return {
        "ticker": normalized_ticker,
        "name": normalized_name,
        "currency": normalized_currency,
        "amount": float(value),
        "serial": serial,
        "date": _format_date(serial),
        "year": int((_date_from_value(date_value) or (BASE_DATE + timedelta(days=int(serial or 0)))).year) if serial else None,
        "kind": clean_text(kind),
        "_source": clean_text(source),
        "_source_row": clean_text(source_row),
    }


def _headers(row) -> dict[str, int]:
    return {clean_text(value): index for index, value in enumerate(row) if clean_text(value)}


def load_workbook_dividend_events(core, workbook_path: Path) -> list[dict[str, object]]:
    load_workbook = _load_workbook()
    if load_workbook is None or not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if DIVIDEND_SHEET_NAME not in workbook.sheetnames:
            return []
        sheet = workbook[DIVIDEND_SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        header = _headers(next(rows, ()))
        required = {"日期", "代码", "事件", "金额"}
        if not required.issubset(header):
            return []
        events: list[dict[str, object]] = []
        for row_number, row in enumerate(rows, start=2):
            kind = clean_text(row[header["事件"]] if header["事件"] < len(row) else "")
            if kind not in DIVIDEND_EVENTS:
                continue
            event = _event(
                core,
                date_value=row[header["日期"]] if header["日期"] < len(row) else "",
                ticker=row[header["代码"]] if header["代码"] < len(row) else "",
                name=row[header.get("名称", -1)] if 0 <= header.get("名称", -1) < len(row) else "",
                kind=kind,
                amount=row[header["金额"]] if header["金额"] < len(row) else "",
                currency=row[header.get("币种", -1)] if 0 <= header.get("币种", -1) < len(row) else "",
                source=row[header.get("来源", -1)] if 0 <= header.get("来源", -1) < len(row) else "Trade Tracker.xlsx",
                source_row=row[header.get("原始行", -1)] if 0 <= header.get("原始行", -1) < len(row) else row_number,
            )
            if event:
                events.append(event)
        return events
    finally:
        workbook.close()


def workbook_has_dividend_sheet(workbook_path: Path) -> bool:
    load_workbook = _load_workbook()
    if load_workbook is None or not workbook_path.exists():
        return False
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return DIVIDEND_SHEET_NAME in workbook.sheetnames
    finally:
        workbook.close()


def load_history_dividend_events(core) -> list[dict[str, object]]:
    load_workbook = _load_workbook()
    if load_workbook is None:
        return []
    events: list[dict[str, object]] = []
    for path in iter_history_workbooks(core):
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            if "交易记录" not in workbook.sheetnames:
                continue
            sheet = workbook["交易记录"]
            rows = sheet.iter_rows(values_only=True)
            header = _headers(next(rows, ()))
            if not {"成交日期", "代码", "名称", "交易类别", "发生金额"}.issubset(header):
                continue
            for row_number, row in enumerate(rows, start=2):
                kind = clean_text(row[header["交易类别"]] if header["交易类别"] < len(row) else "")
                if kind not in DIVIDEND_EVENTS:
                    continue
                event = _event(
                    core,
                    date_value=row[header["成交日期"]] if header["成交日期"] < len(row) else "",
                    ticker=row[header["代码"]] if header["代码"] < len(row) else "",
                    name=row[header["名称"]] if header["名称"] < len(row) else "",
                    kind=kind,
                    amount=row[header["发生金额"]] if header["发生金额"] < len(row) else "",
                    currency="",
                    source=Path(path).name,
                    source_row=row_number,
                )
                if event:
                    events.append(event)
        finally:
            workbook.close()
    return events


def load_dividend_events(core, workbook_path: Path, original_loader=None) -> list[dict[str, object]]:
    if workbook_has_dividend_sheet(workbook_path):
        events = load_workbook_dividend_events(core, workbook_path)
    else:
        events = load_history_dividend_events(core)
    if not events and original_loader is not None:
        try:
            events.extend(original_loader() or [])
        except Exception:
            pass

    deduped: dict[tuple[object, ...], dict[str, object]] = {}
    for event in events:
        deduped.setdefault(_source_key(event), event)
    return [_public_event(event) for event in deduped.values()]


def build_dividend_income_maps(data: dict[str, object]) -> dict[tuple[str, str], float]:
    adjustments: dict[tuple[str, str], float] = {}
    for row in data.get("stock_summary", []) or []:
        ticker = clean_text(row.get("ticker"))
        currency = clean_text(row.get("currency"))
        if not ticker or not currency:
            continue
        dividend = parse_float(str(row.get("dividend") or "").replace(",", "").replace(currency, ""))
        if dividend is None or abs(dividend) < 1e-9:
            continue
        key = (ticker, currency)
        adjustments[key] = adjustments.get(key, 0.0) + float(dividend)
    return adjustments
