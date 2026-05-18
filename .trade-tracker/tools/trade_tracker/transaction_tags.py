from __future__ import annotations

import html
import json
import re
from pathlib import Path

from . import state
from .utils import cell_raw, cell_text, clean_text


TAG_HEADER_ALIASES = {"标签", "交易标签", "Tag", "Tags", "tag", "tags"}
TAG_SPLIT_RE = re.compile(r"[,，;；|、\n\r\t ]+")
DETAILS_PATTERN = re.compile(
    r'<details class="dashboard-section section-collapsible"(?: [^>]*)?>.*?</details>',
    re.S,
)
TIMELINE_TABLE_PATTERN = re.compile(
    r'(<h2 class="section-title">交易时间线</h2>.*?<table\b[^>]*class="[^"]*\bsummary-table\b[^"]*"[^>]*>)'
    r"(.*?)"
    r"(</table>)",
    re.S,
)


def parse_tags(value: object) -> list[str]:
    text = clean_text(value)
    if not text or text in {"-", "--"}:
        return []
    tags: list[str] = []
    seen: set[str] = set()
    for part in TAG_SPLIT_RE.split(text):
        tag = clean_text(part).strip("#＃")
        if not tag or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)
    return tags


def find_tag_column(workbook_path: Path) -> int | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if "交易记录" not in workbook.sheetnames:
            return None
        sheet = workbook["交易记录"]
        for column in range(1, sheet.max_column + 1):
            label = clean_text(sheet.cell(1, column).value)
            if label in TAG_HEADER_ALIASES:
                return column
        return None
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def load_transaction_tags(workbook_path: Path) -> dict[int, list[str]]:
    column = find_tag_column(workbook_path)
    state.TRANSACTION_TAG_COLUMN = column
    if column is None:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        if "交易记录" not in workbook.sheetnames:
            return {}
        sheet = workbook["交易记录"]
        by_row: dict[int, list[str]] = {}
        for row_number in range(2, sheet.max_row + 1):
            tags = parse_tags(sheet.cell(row_number, column).value)
            if tags:
                by_row[row_number] = tags
        return by_row
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def transaction_tags_for_row(row_number: int, cells: dict[int, object] | None = None) -> list[str]:
    tags = state.TRANSACTION_TAGS_BY_ROW.get(row_number)
    if tags:
        return list(tags)
    column = state.TRANSACTION_TAG_COLUMN
    if not cells or column is None:
        return []
    return parse_tags(cell_raw(cells, column))


def tag_counts(tags_by_row: dict[int, list[str]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for tags in tags_by_row.values():
        for tag in tags:
            counts[tag] = counts.get(tag, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0].lower())))


def build_transaction_tags_payload(tags_by_row: dict[int, list[str]]) -> dict[str, object]:
    counts = tag_counts(tags_by_row)
    return {
        "version": 1,
        "taggedRows": len(tags_by_row),
        "tags": [{"tag": tag, "count": count} for tag, count in counts.items()],
    }


def initialize_transaction_tags(workbook_path: Path) -> dict[str, object]:
    tags_by_row = load_transaction_tags(workbook_path)
    state.TRANSACTION_TAGS_BY_ROW = tags_by_row
    state.TRANSACTION_TAGS_PAYLOAD = build_transaction_tags_payload(tags_by_row)
    return state.TRANSACTION_TAGS_PAYLOAD


def tags_for_rows(rows: list[tuple[int, dict[int, object]]]) -> list[list[str]]:
    return [transaction_tags_for_row(row_number, cells) for row_number, cells in rows]


def tag_cell(tags: list[str]) -> str:
    if not tags:
        return '<td class="text transaction-tag-cell">-</td>'
    chips = "".join(f'<span class="transaction-tag-chip">{html.escape(tag)}</span>' for tag in tags)
    return f'<td class="text transaction-tag-cell">{chips}</td>'


def row_tag_attr(tags: list[str]) -> str:
    if not tags:
        return ' data-transaction-tags="" data-transaction-tag-empty="true"'
    encoded = "|".join(tags)
    return f' data-transaction-tags="{html.escape(encoded, quote=True)}"'


def render_timeline_tag_toolbar(tags_by_row: dict[int, list[str]], row_count: int) -> str:
    counts = tag_counts(tags_by_row)
    if not counts:
        return """
            <div class="transaction-tag-toolbar is-empty" data-transaction-tag-toolbar>
              <span class="filter-label">标签</span>
              <span class="transaction-tag-hint">在交易记录新增“标签”列后，这里会自动出现筛选。</span>
            </div>
        """
    buttons = [
        f'<button type="button" class="transaction-tag-filter is-active" data-tag-filter="__all__">全部 <small>{row_count}</small></button>',
        f'<button type="button" class="transaction-tag-filter" data-tag-filter="__none__">未打标签</button>',
    ]
    for tag, count in counts.items():
        buttons.append(
            f'<button type="button" class="transaction-tag-filter" data-tag-filter="{html.escape(tag, quote=True)}">'
            f"{html.escape(tag)} <small>{count}</small></button>"
        )
    return f"""
            <div class="transaction-tag-toolbar" data-transaction-tag-toolbar>
              <span class="filter-label">标签</span>
              <div class="transaction-tag-filters">{''.join(buttons)}</div>
              <span class="transaction-tag-status" data-transaction-tag-status></span>
            </div>
        """


def render_timeline_tag_script() -> str:
    return """
        <script>
        (function setupTransactionTagTimeline() {
          const toolbar = document.querySelector('[data-transaction-tag-toolbar]');
          if (!toolbar || toolbar.dataset.ready === '1') return;
          toolbar.dataset.ready = '1';
          const section = toolbar.closest('details.dashboard-section');
          const table = section ? section.querySelector('table.summary-table') : null;
          const status = toolbar.querySelector('[data-transaction-tag-status]');
          if (!table) return;
          const rows = Array.from(table.querySelectorAll('tbody tr')).filter((row) => row.children.length > 1);
          const buttons = Array.from(toolbar.querySelectorAll('[data-tag-filter]'));
          if (!buttons.length) return;

          function rowTags(row) {
            return String(row.dataset.transactionTags || '').split('|').filter(Boolean);
          }

          function applyFilter(activeTag) {
            let visible = 0;
            rows.forEach((row) => {
              const tags = rowTags(row);
              const show = activeTag === '__all__'
                || (activeTag === '__none__' && !tags.length)
                || tags.includes(activeTag);
              row.style.display = show ? '' : 'none';
              if (show) visible += 1;
            });
            buttons.forEach((button) => {
              const active = button.dataset.tagFilter === activeTag;
              button.classList.toggle('is-active', active);
              button.setAttribute('aria-pressed', active ? 'true' : 'false');
            });
            if (status) status.textContent = `${visible} / ${rows.length} 笔`;
            if (typeof updateSummaryTables === 'function') updateSummaryTables();
            if (typeof balanceSummaryTableWidths === 'function') requestAnimationFrame(balanceSummaryTableWidths);
          }

          toolbar.addEventListener('click', (event) => {
            if (!(event.target instanceof Element)) return;
            const button = event.target.closest('[data-tag-filter]');
            if (!button) return;
            applyFilter(button.dataset.tagFilter || '__all__');
          });
          applyFilter('__all__');
        })();
        </script>
    """


def insert_tag_column_into_timeline_table(table_html: str, row_tags: list[list[str]]) -> str:
    if "data-transaction-tags-ready" in table_html or ">标签</th>" in table_html:
        return table_html
    header_match = re.search(r"(<thead>\s*<tr>)(.*?)(</tr>\s*</thead>)", table_html, re.S)
    if not header_match:
        return table_html
    headers = re.findall(r"<th\b[^>]*>.*?</th>", header_match.group(2), re.S)
    labels = [cell_text(cell) for cell in headers]
    if "事件" not in labels:
        return table_html
    insert_at = labels.index("事件") + 1
    headers.insert(insert_at, '<th class="text" data-sort-type="text">标签</th>')
    updated = table_html[: header_match.start(2)] + "".join(headers) + table_html[header_match.end(2) :]
    updated = updated.replace("<table ", '<table data-transaction-tags-ready="1" ', 1)
    row_index = -1

    def insert_row(match: re.Match[str]) -> str:
        nonlocal row_index
        row_prefix, row_body, row_suffix = match.group(1), match.group(2), match.group(3)
        cells = re.findall(r"<td\b[^>]*>.*?</td>", row_body, re.S)
        if len(cells) != len(labels):
            return match.group(0)
        row_index += 1
        tags = row_tags[row_index] if row_index < len(row_tags) else []
        cells.insert(insert_at, tag_cell(tags))
        prefix = row_prefix[:-1] + row_tag_attr(tags) + ">" if row_prefix.endswith(">") else row_prefix
        return prefix + "".join(cells) + row_suffix

    return re.sub(r"(<tr\b[^>]*>)(.*?)(</tr>)", insert_row, updated, flags=re.S)


def insert_transaction_tags(html_text: str, rows: list[tuple[int, dict[int, object]]]) -> str:
    if "data-transaction-tags-ready" in html_text:
        return html_text
    row_tags = tags_for_rows(rows)
    tags_by_row = {
        row_number: tags
        for (row_number, _cells), tags in zip(rows, row_tags)
        if tags
    }

    def update_table(match: re.Match[str]) -> str:
        table_html = match.group(1) + match.group(2) + match.group(3)
        return insert_tag_column_into_timeline_table(table_html, row_tags)

    updated = TIMELINE_TABLE_PATTERN.sub(update_table, html_text, count=1)
    if updated == html_text:
        return html_text
    section_start = updated.find('<h2 class="section-title">交易时间线</h2>')
    if section_start < 0:
        return updated
    body_start = updated.find('<div class="summary-wrap">', section_start)
    if body_start < 0:
        return updated
    toolbar = render_timeline_tag_toolbar(tags_by_row, len(row_tags))
    script = render_timeline_tag_script()
    updated = updated[:body_start] + toolbar + "\n" + updated[body_start:]
    details_end = updated.find("</details>", body_start)
    if details_end >= 0:
        updated = updated[:details_end] + script + "\n" + updated[details_end:]
    return updated


def tags_json_for_script(payload: object) -> str:
    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return text.replace("&", "\\u0026").replace("<", "\\u003c").replace(">", "\\u003e")
