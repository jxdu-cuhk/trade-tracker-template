from __future__ import annotations

from datetime import date
from pathlib import Path

from .market_data import display_currency_label
from .utils import clean_text, date_key, format_date, parse_float


def build_last_clear_date_map(core, workbook_path: Path) -> dict[tuple[str, str], str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        if "交易记录" not in workbook.sheetnames:
            return {}
        sheet = workbook["交易记录"]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean_text(cell) for cell in next(rows)]
        except StopIteration:
            return {}
        header_map = {header: index for index, header in enumerate(headers) if header}
        required = ["类型", "开仓", "平仓", "代码", "事件", "数量", "币种"]
        if any(label not in header_map for label in required):
            return {}

        events_by_key: dict[tuple[str, str], list[tuple[object, int, float]]] = {}
        for row in rows:
            event = clean_text(row[header_map["事件"]] if header_map["事件"] < len(row) else "")
            if event != "现股":
                continue
            code = row[header_map["代码"]] if header_map["代码"] < len(row) else None
            currency = core.normalize_currency(row[header_map["币种"]] if header_map["币种"] < len(row) else "")
            ticker = core.normalize_ticker(code, currency)
            quantity = parse_float(row[header_map["数量"]] if header_map["数量"] < len(row) else None)
            open_date = date_key(row[header_map["开仓"]] if header_map["开仓"] < len(row) else None)
            close_date = date_key(row[header_map["平仓"]] if header_map["平仓"] < len(row) else None)
            if not ticker or not currency or quantity in (None, 0) or not open_date:
                continue
            trade_type = clean_text(row[header_map["类型"]] if header_map["类型"] < len(row) else "")
            signed_quantity = -abs(quantity) if trade_type in {"卖出", "卖空"} else quantity
            key = (ticker, currency)
            events_by_key.setdefault(key, []).append((open_date, 0, signed_quantity))
            if close_date:
                events_by_key[key].append((close_date, 1, -signed_quantity))

        last_clear_dates: dict[tuple[str, str], str] = {}
        for key, events in events_by_key.items():
            position = 0.0
            for event_date, _, delta in sorted(events, key=lambda item: (item[0], item[1])):
                position += delta
                if abs(position) < 0.000001:
                    last_clear_dates[key] = format_date(event_date)
        return last_clear_dates
    finally:
        workbook.close()


def build_holding_days_map(core, workbook_path: Path) -> dict[tuple[str, str], str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        if "交易记录" not in workbook.sheetnames:
            return {}
        sheet = workbook["交易记录"]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean_text(cell) for cell in next(rows)]
        except StopIteration:
            return {}
        header_map = {header: index for index, header in enumerate(headers) if header}
        required = ["开仓", "平仓", "代码", "事件", "数量", "币种"]
        if any(label not in header_map for label in required):
            return {}

        today = date.today()
        first_open_by_key: dict[tuple[str, str], date] = {}
        for row in rows:
            event = clean_text(row[header_map["事件"]] if header_map["事件"] < len(row) else "")
            if event != "现股":
                continue
            if date_key(row[header_map["平仓"]] if header_map["平仓"] < len(row) else None):
                continue
            quantity = parse_float(row[header_map["数量"]] if header_map["数量"] < len(row) else None)
            open_date = date_key(row[header_map["开仓"]] if header_map["开仓"] < len(row) else None)
            currency = core.normalize_currency(row[header_map["币种"]] if header_map["币种"] < len(row) else "")
            ticker = core.normalize_ticker(row[header_map["代码"]] if header_map["代码"] < len(row) else None, currency)
            if not ticker or not currency or quantity in (None, 0) or not open_date:
                continue
            key = (ticker, currency)
            if key not in first_open_by_key or open_date < first_open_by_key[key]:
                first_open_by_key[key] = open_date

        return {
            key: str(max((today - open_date).days, 0))
            for key, open_date in first_open_by_key.items()
        }
    finally:
        workbook.close()


def build_summary_holding_days_maps(
    core,
    workbook_path: Path,
) -> tuple[dict[tuple[str, str], int], dict[tuple[str, str, str], int]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}, {}

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return {}, {}

    try:
        if "交易记录" not in workbook.sheetnames:
            return {}, {}
        sheet = workbook["交易记录"]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean_text(cell) for cell in next(rows)]
        except StopIteration:
            return {}, {}
        header_map = {header: index for index, header in enumerate(headers) if header}
        required = ["开仓", "平仓", "代码", "事件", "数量", "币种"]
        if any(label not in header_map for label in required):
            return {}, {}

        today = date.today()
        events_by_key: dict[tuple[str, str], list[tuple[date, int, float]]] = {}
        for row in rows:
            event = clean_text(row[header_map["事件"]] if header_map["事件"] < len(row) else "")
            if event != "现股":
                continue
            quantity = parse_float(row[header_map["数量"]] if header_map["数量"] < len(row) else None)
            open_date = date_key(row[header_map["开仓"]] if header_map["开仓"] < len(row) else None)
            close_date = date_key(row[header_map["平仓"]] if header_map["平仓"] < len(row) else None)
            currency = core.normalize_currency(row[header_map["币种"]] if header_map["币种"] < len(row) else "")
            ticker = core.normalize_ticker(row[header_map["代码"]] if header_map["代码"] < len(row) else None, currency)
            if not ticker or not currency or quantity in (None, 0) or not open_date:
                continue
            currency_label = display_currency_label(core, currency)
            key = (ticker, currency_label)
            type_index = header_map.get("类型")
            trade_type = clean_text(row[type_index] if type_index is not None and type_index < len(row) else "")
            signed_quantity = -abs(quantity) if trade_type in {"卖出", "卖空"} else quantity
            events_by_key.setdefault(key, []).append((open_date, 0, signed_quantity))
            if close_date:
                events_by_key[key].append((close_date, 1, -signed_quantity))

        total_days: dict[tuple[str, str], int] = {}
        annual_days: dict[tuple[str, str, str], int] = {}

        def add_period(key: tuple[str, str], start: date, end: date) -> None:
            days = max((end - start).days, 1)
            total_days[key] = total_days.get(key, 0) + days
            ticker, currency_label = key
            if end <= start:
                annual_key = (ticker, currency_label, str(start.year))
                annual_days[annual_key] = annual_days.get(annual_key, 0) + 1
                return
            for year in range(start.year, end.year + 1):
                year_start = date(year, 1, 1)
                year_end = date(year + 1, 1, 1)
                segment_start = max(start, year_start)
                segment_end = min(end, year_end)
                if segment_end <= segment_start:
                    continue
                annual_key = (ticker, currency_label, str(year))
                annual_days[annual_key] = annual_days.get(annual_key, 0) + (segment_end - segment_start).days

        for key, events in events_by_key.items():
            position = 0.0
            active_start = None
            for event_date, _order, delta in sorted(events, key=lambda item: (item[0], item[1])):
                was_flat = abs(position) < 0.000001
                position += delta
                is_flat = abs(position) < 0.000001
                if was_flat and not is_flat:
                    active_start = event_date
                elif not was_flat and is_flat and active_start is not None:
                    add_period(key, active_start, event_date)
                    active_start = None
            if active_start is not None and abs(position) >= 0.000001:
                add_period(key, active_start, today)
        return total_days, annual_days
    finally:
        workbook.close()
