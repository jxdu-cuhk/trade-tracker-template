from __future__ import annotations

import csv
import json
from pathlib import Path

from .runtime import HISTORY_DIR, NAME_CACHE_PATH
from .utils import clean_name, clean_text, normalize_source_code

_NAME_CACHE: dict[str, str] | None = None
_NAME_CACHE_DIRTY = False


def name_cache_key(core, ticker, currency) -> str:
    normalized_currency = core.normalize_currency(currency)
    normalized_ticker = core.normalize_ticker(ticker, normalized_currency)
    if not normalized_currency:
        normalized_currency = "HKD" if normalized_ticker.isdigit() and len(normalized_ticker) == 5 else "CNY"
    return f"{normalized_currency}:{normalized_ticker}"


def load_name_cache() -> dict[str, str]:
    global _NAME_CACHE
    if _NAME_CACHE is not None:
        return _NAME_CACHE
    names: dict[str, str] = {}
    if NAME_CACHE_PATH.exists():
        try:
            payload = json.loads(NAME_CACHE_PATH.read_text(encoding="utf-8"))
            raw_names = payload.get("names", payload) if isinstance(payload, dict) else {}
            if isinstance(raw_names, dict):
                for key, value in raw_names.items():
                    name = clean_name(value)
                    if name:
                        names[str(key)] = name
        except (OSError, ValueError, TypeError):
            names = {}
    _NAME_CACHE = names
    return _NAME_CACHE


def cache_name(core, ticker, currency, name) -> str:
    global _NAME_CACHE_DIRTY
    cleaned = clean_name(name)
    if not cleaned:
        return ""
    key = name_cache_key(core, ticker, currency)
    names = load_name_cache()
    if names.get(key) != cleaned:
        names[key] = cleaned
        _NAME_CACHE_DIRTY = True
    return cleaned


def save_name_cache() -> None:
    global _NAME_CACHE_DIRTY
    if not _NAME_CACHE_DIRTY or _NAME_CACHE is None:
        return
    payload = {"version": 1, "names": dict(sorted(_NAME_CACHE.items()))}
    NAME_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = NAME_CACHE_PATH.with_suffix(".json.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp_path.replace(NAME_CACHE_PATH)
    _NAME_CACHE_DIRTY = False


def put_source_name(mapping: dict[str, str], code, name) -> None:
    normalized_code = normalize_source_code(code)
    normalized_name = clean_name(name)
    if not normalized_code or not normalized_name:
        return
    keys = {normalized_code}
    if normalized_code.isdigit():
        keys.add(normalized_code.zfill(5))
        keys.add(normalized_code.zfill(6))
    for key in keys:
        mapping.setdefault(key, normalized_name)


def iter_history_workbooks(core):
    seen: set[Path] = set()
    for path in getattr(core, "NAME_SOURCE_FILES", []):
        path = Path(path)
        if path.exists() and path.suffix.lower() in {".xlsx", ".xlsm"}:
            seen.add(path.resolve())
            yield path
    if HISTORY_DIR.exists():
        for path in sorted(HISTORY_DIR.glob("*.xlsx")):
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                yield path


def iter_history_csvs(core):
    seen: set[Path] = set()
    for path in getattr(core, "CSV_NAME_SOURCES", []):
        path = Path(path)
        if path.exists() and path.suffix.lower() == ".csv":
            seen.add(path.resolve())
            yield path
    if HISTORY_DIR.exists():
        for path in sorted(HISTORY_DIR.glob("*.csv")):
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                yield path


def load_workbook_name_map(core) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for code, name in getattr(core, "LOCAL_FALLBACK_NAMES", {}).items():
        put_source_name(mapping, code, name)

    try:
        from openpyxl import load_workbook
    except ImportError:
        load_workbook = None

    if load_workbook is not None:
        for path in iter_history_workbooks(core):
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            try:
                for sheet in workbook.worksheets:
                    header_index = None
                    code_index = None
                    name_index = None
                    rows = sheet.iter_rows(values_only=True)
                    for row_number, row in enumerate(rows, start=1):
                        labels = [clean_text(cell) for cell in row]
                        if "代码" in labels and "名称" in labels:
                            header_index = row_number
                            code_index = labels.index("代码")
                            name_index = labels.index("名称")
                            break
                        if row_number >= 20:
                            break
                    if header_index is None or code_index is None or name_index is None:
                        continue
                    for row in sheet.iter_rows(min_row=header_index + 1, values_only=True):
                        if max(code_index, name_index) >= len(row):
                            continue
                        put_source_name(mapping, row[code_index], row[name_index])
            finally:
                workbook.close()

    for path in iter_history_csvs(core):
        for encoding in ("utf-8-sig", "gb18030", "utf-16"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    if not reader.fieldnames or "代码" not in reader.fieldnames or "名称" not in reader.fieldnames:
                        continue
                    for row in reader:
                        put_source_name(mapping, row.get("代码"), row.get("名称"))
                break
            except (OSError, UnicodeError, csv.Error):
                continue
    return mapping
