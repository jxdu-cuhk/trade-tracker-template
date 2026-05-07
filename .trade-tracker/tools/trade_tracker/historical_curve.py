from __future__ import annotations

import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from dataclasses import replace
from datetime import date, datetime, timedelta, timezone
from urllib.error import URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from .market_data import display_currency_label, infer_secid, tencent_symbol
from .runtime import APP_DIR, emit_progress
from .settings import OPTION_EVENTS
from .utils import cell_raw, clean_text, core_trade_type, excel_serial_to_date, parse_display_number, parse_float, raw_number, raw_text_value


STOCK_EVENTS = {"现股", "Stock", "stock", "STOCK"}
HISTORY_CACHE_PATH = APP_DIR / "tools" / "cache" / "security_history.json"
HISTORY_SOURCE_DIR = APP_DIR / "history"
HISTORY_HTTP_TIMEOUT = 8
HISTORY_WORKERS = 5
EPSILON = 0.000001


@dataclass(frozen=True)
class SecurityHistoryPoint:
    iso: str
    close: float


@dataclass(frozen=True)
class StockCurveLot:
    ticker: str
    currency: str
    currency_raw: str
    open_date: date
    close_date: date | None
    quantity: float
    open_price: float
    close_price: float | None
    fee: float
    entry_fee: float | None
    capital: float
    realized_pnl: float | None
    source: str = ""

    @property
    def is_short(self) -> bool:
        return self.quantity < 0

    @property
    def end_date(self) -> date:
        return self.close_date or date.today()


@dataclass(frozen=True)
class RealizedCurveEvent:
    event_date: date
    currency: str
    pnl: float
    ticker: str = ""


def date_label(day: date) -> str:
    return day.strftime("%Y/%m/%d")


def excel_serial(day: date) -> float:
    return float((day - date(1899, 12, 30)).days)


def date_range(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def load_history_cache() -> dict[str, object]:
    try:
        payload = json.loads(HISTORY_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    securities = payload.get("securities")
    if not isinstance(securities, dict):
        payload["securities"] = {}
    payload["version"] = 1
    return payload


def save_history_cache(payload: dict[str, object]) -> None:
    try:
        HISTORY_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        HISTORY_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        return


def history_cache_key(ticker: str, currency: str) -> str:
    return f"{clean_text(currency).upper()}|{clean_text(ticker).upper()}"


def point_from_cache(item: object) -> SecurityHistoryPoint | None:
    if not isinstance(item, dict):
        return None
    iso = clean_text(item.get("iso"))
    close = parse_float(item.get("close"))
    if not iso or close is None or close <= 0:
        return None
    return SecurityHistoryPoint(iso, float(close))


def sorted_history_points(points: list[SecurityHistoryPoint]) -> list[SecurityHistoryPoint]:
    by_iso = {point.iso: point for point in points if point.iso and point.close > 0}
    return [by_iso[iso] for iso in sorted(by_iso)]


def slice_history_points(points: list[SecurityHistoryPoint], start: date, end: date) -> list[SecurityHistoryPoint]:
    start_iso = start.isoformat()
    end_iso = end.isoformat()
    return [point for point in points if start_iso <= point.iso <= end_iso]


def entry_points(entry: object) -> list[SecurityHistoryPoint]:
    if not isinstance(entry, dict):
        return []
    points = entry.get("points")
    if not isinstance(points, list):
        return []
    parsed = [point for item in points if (point := point_from_cache(item))]
    return sorted_history_points(parsed)


def entry_was_checked_today(entry: object) -> bool:
    if not isinstance(entry, dict):
        return False
    try:
        fetched_at = date.fromisoformat(clean_text(entry.get("fetched_at")))
    except ValueError:
        return False
    return fetched_at >= date.today()


def entry_covers(points: list[SecurityHistoryPoint], start: date, end: date, entry: object = None) -> bool:
    if not points:
        return entry_was_checked_today(entry)
    if points[0].iso > start.isoformat():
        return entry_was_checked_today(entry)
    if points[-1].iso >= end.isoformat():
        return True
    return entry_was_checked_today(entry)


def merge_cache_points(existing: list[SecurityHistoryPoint], new_points: list[SecurityHistoryPoint]) -> list[SecurityHistoryPoint]:
    return sorted_history_points([*existing, *new_points])


def serialize_history_points(points: list[SecurityHistoryPoint]) -> list[dict[str, object]]:
    return [{"iso": point.iso, "close": point.close} for point in points]


def parse_eastmoney_kline_rows(rows: object) -> list[SecurityHistoryPoint]:
    points: list[SecurityHistoryPoint] = []
    if not isinstance(rows, list):
        return points
    for row in rows:
        parts = clean_text(row).split(",")
        if len(parts) < 3:
            continue
        iso = clean_text(parts[0]).replace("/", "-")
        close = parse_float(parts[2])
        if iso and close is not None and close > 0:
            points.append(SecurityHistoryPoint(iso, float(close)))
    return sorted_history_points(points)


def parse_tencent_kline_rows(rows: object) -> list[SecurityHistoryPoint]:
    points: list[SecurityHistoryPoint] = []
    if not isinstance(rows, list):
        return points
    for row in rows:
        if not isinstance(row, list) or len(row) < 3:
            continue
        iso = clean_text(row[0]).replace("/", "-")
        close = parse_float(row[2])
        if iso and close is not None and close > 0:
            points.append(SecurityHistoryPoint(iso, float(close)))
    return sorted_history_points(points)


def fetch_tencent_history_points(symbol: str, start: date, end: date) -> list[SecurityHistoryPoint]:
    symbol = clean_text(symbol)
    if not symbol:
        return []
    day_count = max((end - start).days + 10, 120)
    params = {
        "param": f"{symbol},day,{start.isoformat()},{end.isoformat()},{day_count},",
    }
    url = "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?" + urlencode(params)
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=HISTORY_HTTP_TIMEOUT) as response:
            payload = json.loads(response.read().decode("utf-8", "ignore"))
    except (TimeoutError, URLError, OSError, ValueError):
        return []
    data = (payload.get("data") or {}).get(symbol) if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        return []
    rows = data.get("day") or data.get("qfqday") or []
    return parse_tencent_kline_rows(rows)


def fetch_eastmoney_history_points(secid: str, start: date, end: date) -> list[SecurityHistoryPoint]:
    params = {
        "secid": secid,
        "fields1": "f1,f2,f3,f4,f5,f6",
        "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61",
        "klt": "101",
        "fqt": "0",
        "beg": start.strftime("%Y%m%d"),
        "end": end.strftime("%Y%m%d"),
    }
    url = "https://push2his.eastmoney.com/api/qt/stock/kline/get?" + urlencode(params)
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=HISTORY_HTTP_TIMEOUT) as response:
            payload = json.loads(response.read().decode("utf-8", "ignore"))
    except (TimeoutError, URLError, OSError, ValueError):
        return []
    rows = ((payload.get("data") or {}).get("klines") or []) if isinstance(payload, dict) else []
    return parse_eastmoney_kline_rows(rows)


def fetch_yahoo_history_points(symbol: str, start: date, end: date) -> list[SecurityHistoryPoint]:
    period1 = int(datetime(start.year, start.month, start.day, tzinfo=timezone.utc).timestamp())
    period2_day = end + timedelta(days=1)
    period2 = int(datetime(period2_day.year, period2_day.month, period2_day.day, tzinfo=timezone.utc).timestamp())
    url = "https://query1.finance.yahoo.com/v8/finance/chart/" + clean_text(symbol).upper() + "?" + urlencode(
        {
            "period1": str(period1),
            "period2": str(period2),
            "interval": "1d",
            "events": "history",
            "includeAdjustedClose": "true",
        }
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=HISTORY_HTTP_TIMEOUT) as response:
            payload = json.loads(response.read().decode("utf-8", "ignore"))
    except (TimeoutError, URLError, OSError, ValueError):
        return []
    try:
        result = payload["chart"]["result"][0]
        timestamps = result.get("timestamp") or []
        quote = (result.get("indicators") or {}).get("quote", [{}])[0]
        closes = quote.get("close") or []
    except (KeyError, IndexError, TypeError):
        return []
    points: list[SecurityHistoryPoint] = []
    for timestamp, close in zip(timestamps, closes):
        if not isinstance(timestamp, (int, float)) or not isinstance(close, (int, float)) or close <= 0:
            continue
        iso = datetime.fromtimestamp(timestamp, timezone.utc).date().isoformat()
        points.append(SecurityHistoryPoint(iso, float(close)))
    return sorted_history_points(points)


def fetch_security_history_points_online(core, ticker: str, currency: str, start: date, end: date) -> list[SecurityHistoryPoint]:
    try:
        normalized_currency = core.normalize_currency(currency)
    except Exception:
        normalized_currency = clean_text(currency).upper()
    points: list[SecurityHistoryPoint] = []
    symbol = tencent_symbol(core, ticker, normalized_currency)
    if symbol:
        points = fetch_tencent_history_points(symbol, start, end)
        if points and (len(points) >= 2 or (end - start).days <= 7):
            return points
    if normalized_currency == "USD":
        fallback = fetch_yahoo_history_points(ticker, start, end)
        return fallback or points
    secid = infer_secid(core, ticker, normalized_currency)
    if not secid:
        return points
    fallback = fetch_eastmoney_history_points(secid, start, end)
    return fallback if len(fallback) > len(points) else (points or fallback)


def fetch_security_history_points(core, ticker: str, currency: str, start: date, end: date) -> list[SecurityHistoryPoint]:
    if not ticker or not currency or end < start:
        return []
    cache = load_history_cache()
    securities = cache.get("securities")
    if not isinstance(securities, dict):
        securities = {}
        cache["securities"] = securities
    key = history_cache_key(ticker, currency)
    entry = securities.get(key)
    cached_points = entry_points(entry)
    if entry_covers(cached_points, start, end, entry):
        return slice_history_points(cached_points, start, end)

    fetch_ranges: list[tuple[date, date]] = []
    if not cached_points:
        fetch_ranges.append((start, end))
    else:
        first_cached = date.fromisoformat(cached_points[0].iso)
        last_cached = date.fromisoformat(cached_points[-1].iso)
        if start < first_cached:
            fetch_ranges.append((start, first_cached - timedelta(days=1)))
        if end > last_cached:
            fetch_ranges.append((last_cached + timedelta(days=1), end))

    fetched: list[SecurityHistoryPoint] = []
    for range_start, range_end in fetch_ranges:
        if range_end < range_start:
            continue
        fetched.extend(fetch_security_history_points_online(core, ticker, currency, range_start, range_end))

    if fetched or cached_points:
        merged = merge_cache_points(cached_points, fetched)
        securities[key] = {
            "ticker": clean_text(ticker),
            "currency": clean_text(currency),
            "fetched_at": date.today().isoformat(),
            "points": serialize_history_points(merged),
        }
        save_history_cache(cache)
        if fetched:
            return slice_history_points(merged, start, end)
    return slice_history_points(cached_points, start, end)


def normalize_currency_label(core, currency: object) -> tuple[str, str]:
    try:
        normalized = core.normalize_currency(currency)
    except Exception:
        normalized = clean_text(currency)
    return display_currency_label(core, normalized), clean_text(normalized)


def imported_trade_source_from_note(note: object) -> str:
    text = clean_text(note)
    prefix = "导入自 "
    suffix = " 成交记录"
    if not text.startswith(prefix) or suffix not in text:
        return ""
    return clean_text(text[len(prefix) : text.find(suffix)])


def stock_lot_from_row(core, cells: dict[int, object]) -> StockCurveLot | None:
    event = raw_text_value(core, cells, 6)
    if event not in STOCK_EVENTS:
        return None
    open_date = excel_serial_to_date(cell_raw(cells, 2))
    if open_date is None:
        return None
    quantity = raw_number(cells, 8)
    open_price = raw_number(cells, 9)
    if quantity in (None, 0) or open_price is None or open_price <= 0:
        return None
    currency, currency_raw = normalize_currency_label(core, cell_raw(cells, 20))
    try:
        ticker = clean_text(core.normalize_ticker(cell_raw(cells, 5), currency_raw))
    except Exception:
        ticker = clean_text(cell_raw(cells, 5)).upper()
    if not ticker or not currency:
        return None
    trade_type = core_trade_type(raw_text_value(core, cells, 1))
    signed_quantity = -abs(float(quantity)) if trade_type == "sell" else abs(float(quantity))
    close_date = excel_serial_to_date(cell_raw(cells, 4))
    close_price = raw_number(cells, 10)
    fee = abs(raw_number(cells, 11) or 0.0)
    capital = raw_number(cells, 12)
    if capital is None or abs(capital) <= EPSILON:
        capital = abs(float(quantity) * open_price) + fee
    realized_pnl = None
    if close_date is not None or close_price is not None:
        try:
            metrics = core.compute_row_metrics(cells)
        except Exception:
            metrics = {}
        if isinstance(metrics, dict):
            realized_pnl = parse_float(metrics.get("pnl"))
        if realized_pnl is None and close_price is not None:
            if signed_quantity < 0:
                realized_pnl = (open_price - close_price) * abs(signed_quantity) - fee
            else:
                realized_pnl = (close_price - open_price) * abs(signed_quantity) - fee
    return StockCurveLot(
        ticker=ticker,
        currency=currency,
        currency_raw=currency_raw,
        open_date=open_date,
        close_date=close_date,
        quantity=signed_quantity,
        open_price=float(open_price),
        close_price=float(close_price) if close_price is not None else None,
        fee=fee,
        entry_fee=fee,
        capital=abs(float(capital)),
        realized_pnl=realized_pnl,
        source=imported_trade_source_from_note(raw_text_value(core, cells, 18)),
    )


def raw_trade_date(value: object) -> date | None:
    return excel_serial_to_date(value)


def raw_trade_side(value: object) -> str:
    text = clean_text(value)
    if "买入" in text or "买券" in text:
        return "buy"
    if "卖出" in text or "卖券" in text:
        return "sell"
    return ""


def source_workbook_path(source: str):
    source = clean_text(source)
    return HISTORY_SOURCE_DIR / f"{source}.xlsx" if source else None


def load_raw_source_closed_lots(core, sources: set[str]) -> list[StockCurveLot]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    all_lots: list[StockCurveLot] = []
    for source in sorted(clean_text(item) for item in sources if clean_text(item)):
        path = source_workbook_path(source)
        if path is None or not path.exists():
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except OSError:
            continue
        try:
            if "交易记录" not in workbook.sheetnames:
                continue
            currency, currency_raw = normalize_currency_label(core, "人民币")
            open_positions: dict[str, list[dict[str, float | date | str]]] = {}
            for row in workbook["交易记录"].iter_rows(min_row=2, values_only=True):
                if len(row) < 10:
                    continue
                trade_date = raw_trade_date(row[0])
                side = raw_trade_side(row[4])
                quantity = parse_float(row[5])
                price = parse_float(row[6])
                amount = parse_float(row[8])
                fee = abs(parse_float(row[9]) or 0.0)
                if trade_date is None or not side or quantity is None or price is None:
                    continue
                quantity = abs(float(quantity))
                price = float(price)
                if quantity <= EPSILON or price <= 0:
                    continue
                try:
                    ticker = clean_text(core.normalize_ticker(row[2], currency_raw))
                except Exception:
                    ticker = clean_text(row[2]).upper()
                if not ticker:
                    continue
                if side == "buy":
                    capital = abs(float(amount)) if amount is not None and abs(float(amount)) > EPSILON else quantity * price
                    open_positions.setdefault(ticker, []).append(
                        {
                            "date": trade_date,
                            "quantity": quantity,
                            "remaining": quantity,
                            "price": price,
                            "fee": fee,
                            "remaining_fee": fee,
                            "capital": capital + fee,
                            "remaining_capital": capital + fee,
                        }
                    )
                    continue

                sell_remaining = quantity
                sell_fee_remaining = fee
                sell_queue = open_positions.get(ticker, [])
                while sell_remaining > EPSILON and sell_queue:
                    buy_lot = sell_queue[0]
                    buy_remaining = float(buy_lot.get("remaining") or 0.0)
                    if buy_remaining <= EPSILON:
                        sell_queue.pop(0)
                        continue
                    matched_qty = min(buy_remaining, sell_remaining)
                    buy_ratio = matched_qty / buy_remaining if buy_remaining > EPSILON else 0.0
                    sell_ratio = matched_qty / sell_remaining if sell_remaining > EPSILON else 0.0
                    buy_fee = float(buy_lot.get("remaining_fee") or 0.0) * buy_ratio
                    sell_fee = sell_fee_remaining * sell_ratio
                    buy_capital = float(buy_lot.get("remaining_capital") or 0.0) * buy_ratio
                    open_price = float(buy_lot["price"])
                    realized_pnl = (price - open_price) * matched_qty - buy_fee - sell_fee
                    all_lots.append(
                        StockCurveLot(
                            ticker=ticker,
                            currency=currency,
                            currency_raw=currency_raw,
                            open_date=buy_lot["date"],
                            close_date=trade_date,
                            quantity=matched_qty,
                            open_price=open_price,
                            close_price=price,
                            fee=buy_fee + sell_fee,
                            entry_fee=buy_fee,
                            capital=buy_capital,
                            realized_pnl=realized_pnl,
                            source=source,
                        )
                    )
                    buy_lot["remaining"] = buy_remaining - matched_qty
                    buy_lot["remaining_fee"] = float(buy_lot.get("remaining_fee") or 0.0) - buy_fee
                    buy_lot["remaining_capital"] = float(buy_lot.get("remaining_capital") or 0.0) - buy_capital
                    sell_remaining -= matched_qty
                    sell_fee_remaining -= sell_fee
                    if float(buy_lot.get("remaining") or 0.0) <= EPSILON:
                        sell_queue.pop(0)
        finally:
            workbook.close()
    return all_lots


def lot_group_key(lot: StockCurveLot) -> tuple[str, str, str]:
    return (clean_text(lot.source), clean_text(lot.ticker), clean_text(lot.currency))


def reconcile_detailed_lot_pnl_to_aggregate(
    detailed_lots: list[StockCurveLot],
    aggregate_lots: list[StockCurveLot],
) -> list[StockCurveLot]:
    """Keep imported fills for dates/capital, but match the workbook P&L totals."""
    if not detailed_lots or not aggregate_lots:
        return detailed_lots

    target_by_group: dict[tuple[str, str, str], float] = {}
    for lot in aggregate_lots:
        if lot.realized_pnl is None:
            continue
        key = lot_group_key(lot)
        target_by_group[key] = target_by_group.get(key, 0.0) + float(lot.realized_pnl)

    by_group: dict[tuple[str, str, str], list[StockCurveLot]] = {}
    for lot in detailed_lots:
        by_group.setdefault(lot_group_key(lot), []).append(lot)

    reconciled: list[StockCurveLot] = []
    for key, lots in by_group.items():
        target = target_by_group.get(key)
        if target is None:
            reconciled.extend(lots)
            continue
        current = sum(float(lot.realized_pnl or 0.0) for lot in lots)
        if abs(current) > EPSILON:
            factor = target / current
            reconciled.extend(replace(lot, realized_pnl=float(lot.realized_pnl or 0.0) * factor) for lot in lots)
            continue
        if not lots:
            continue
        latest_index = max(range(len(lots)), key=lambda index: lots[index].close_date or lots[index].open_date)
        for index, lot in enumerate(lots):
            realized = float(lot.realized_pnl or 0.0)
            if index == latest_index:
                realized += target
            reconciled.append(replace(lot, realized_pnl=realized))
    return reconciled


def non_stock_realized_event_from_row(core, cells: dict[int, object]) -> RealizedCurveEvent | None:
    event = raw_text_value(core, cells, 6)
    if event in STOCK_EVENTS:
        return None
    if event not in OPTION_EVENTS:
        return None
    close_date = excel_serial_to_date(cell_raw(cells, 4))
    if close_date is None and raw_number(cells, 10) is None:
        return None
    try:
        metrics = core.compute_row_metrics(cells)
    except Exception:
        metrics = {}
    pnl = parse_float(metrics.get("pnl") if isinstance(metrics, dict) else None)
    if pnl is None:
        return None
    currency, _currency_raw = normalize_currency_label(core, cell_raw(cells, 20))
    if not currency:
        return None
    try:
        ticker = clean_text(core.normalize_ticker(cell_raw(cells, 5), cell_raw(cells, 20)))
    except Exception:
        ticker = clean_text(cell_raw(cells, 5)).upper()
    return RealizedCurveEvent(close_date or date.today(), currency, float(pnl), ticker)


def build_stock_lots(core, rows: list[tuple[int, dict[int, object]]]) -> list[StockCurveLot]:
    lots = []
    for _row_number, cells in rows:
        lot = stock_lot_from_row(core, cells)
        if lot:
            lots.append(lot)
    aggregate_sources = {lot.source for lot in lots if lot.source and lot.close_date is not None}
    detailed_lots = load_raw_source_closed_lots(core, aggregate_sources)
    detailed_sources = {lot.source for lot in detailed_lots}
    if detailed_sources:
        aggregate_closed_lots = [lot for lot in lots if lot.source in detailed_sources and lot.close_date is not None]
        aggregate_groups = {lot_group_key(lot) for lot in aggregate_closed_lots}
        detailed_lots = [lot for lot in detailed_lots if lot_group_key(lot) in aggregate_groups]
        detailed_lots = reconcile_detailed_lot_pnl_to_aggregate(detailed_lots, aggregate_closed_lots)
        detailed_groups = {lot_group_key(lot) for lot in detailed_lots}
        lots = [lot for lot in lots if not (lot.close_date is not None and lot_group_key(lot) in detailed_groups)]
        lots.extend(detailed_lots)
    return lots


def build_non_stock_realized_events(core, rows: list[tuple[int, dict[int, object]]]) -> list[RealizedCurveEvent]:
    events = []
    for _row_number, cells in rows:
        event = non_stock_realized_event_from_row(core, cells)
        if event:
            events.append(event)
    return events


def build_dividend_realized_events(core) -> list[RealizedCurveEvent]:
    try:
        raw_events = core.load_dividend_events() or []
    except Exception:
        raw_events = []
    events: list[RealizedCurveEvent] = []
    for item in raw_events:
        if not isinstance(item, dict):
            continue
        event_date = excel_serial_to_date(item.get("serial")) or excel_serial_to_date(item.get("date"))
        amount = parse_float(item.get("amount"))
        if event_date is None or amount is None:
            continue
        currency, _currency_raw = normalize_currency_label(core, item.get("currency"))
        if not currency:
            continue
        try:
            ticker = clean_text(core.normalize_ticker(item.get("ticker"), item.get("currency")))
        except Exception:
            ticker = clean_text(item.get("ticker")).upper()
        events.append(RealizedCurveEvent(event_date, currency, float(amount), ticker))
    return events


def summary_realized_targets(data: dict[str, object] | None) -> dict[tuple[str, str], float]:
    if not isinstance(data, dict):
        return {}
    targets: dict[tuple[str, str], float] = {}
    for item in data.get("stock_summary", []) or []:
        if not isinstance(item, dict):
            continue
        ticker = clean_text(item.get("ticker"))
        currency = clean_text(item.get("currency"))
        realized = parse_display_number(item.get("realized_pnl"))
        if ticker and currency and realized is not None:
            targets[(ticker, currency)] = targets.get((ticker, currency), 0.0) + float(realized)
    return targets


def adjust_stock_lot_realized_to_summary(
    lots: list[StockCurveLot],
    data: dict[str, object] | None,
    non_stock_events: list[RealizedCurveEvent],
) -> list[StockCurveLot]:
    targets = summary_realized_targets(data)
    if not targets:
        return lots
    for event in non_stock_events:
        if event.ticker:
            key = (event.ticker, event.currency)
            if key in targets:
                targets[key] -= event.pnl

    by_key: dict[tuple[str, str], list[StockCurveLot]] = {}
    for lot in lots:
        if lot.close_date is not None:
            by_key.setdefault((lot.ticker, lot.currency), []).append(lot)

    replacement_by_id: dict[int, StockCurveLot] = {}
    for key, target in targets.items():
        group = by_key.get(key) or []
        if not group:
            continue
        current = sum(float(lot.realized_pnl or 0.0) for lot in group)
        if abs(current - target) <= 0.01:
            continue
        if abs(current) > EPSILON:
            factor = target / current
            replacement_by_id.update(
                {id(lot): replace(lot, realized_pnl=float(lot.realized_pnl or 0.0) * factor) for lot in group}
            )
        else:
            latest = max(group, key=lambda lot: lot.close_date or lot.open_date)
            replacement_by_id[id(latest)] = replace(latest, realized_pnl=target)

    if not replacement_by_id:
        return lots
    return [replacement_by_id.get(id(lot), lot) for lot in lots]


def residual_realized_events_from_summary(
    core,
    rows: list[tuple[int, dict[int, object]]],
    data: dict[str, object] | None,
    lots: list[StockCurveLot],
    non_stock_events: list[RealizedCurveEvent],
) -> list[RealizedCurveEvent]:
    residuals = summary_realized_targets(data)
    if not residuals:
        return []
    for lot in lots:
        if lot.close_date is not None:
            key = (lot.ticker, lot.currency)
            if key in residuals:
                residuals[key] -= float(lot.realized_pnl or 0.0)
    for event in non_stock_events:
        if event.ticker:
            key = (event.ticker, event.currency)
            if key in residuals:
                residuals[key] -= event.pnl

    latest_date_by_key: dict[tuple[str, str], date] = {}
    for lot in lots:
        key = (lot.ticker, lot.currency)
        event_date = lot.close_date or lot.open_date
        current = latest_date_by_key.get(key)
        latest_date_by_key[key] = max(current, event_date) if current else event_date

    for _row_number, cells in rows:
        try:
            currency, currency_raw = normalize_currency_label(core, cell_raw(cells, 20))
            ticker = clean_text(core.normalize_ticker(cell_raw(cells, 5), currency_raw))
        except Exception:
            currency = clean_text(cell_raw(cells, 20))
            ticker = clean_text(cell_raw(cells, 5)).upper()
        if not ticker or not currency:
            continue
        event_date = excel_serial_to_date(cell_raw(cells, 4)) or excel_serial_to_date(cell_raw(cells, 2))
        if event_date is None:
            continue
        key = (ticker, currency)
        current = latest_date_by_key.get(key)
        latest_date_by_key[key] = max(current, event_date) if current else event_date

    events: list[RealizedCurveEvent] = []
    for (ticker, currency), amount in residuals.items():
        if abs(amount) <= 0.01:
            continue
        event_date = latest_date_by_key.get((ticker, currency)) or date.today()
        events.append(RealizedCurveEvent(event_date, currency, amount, ticker))
    return events


def history_lookup(points: list[SecurityHistoryPoint]) -> dict[date, float]:
    lookup: dict[date, float] = {}
    for point in points:
        try:
            lookup[date.fromisoformat(point.iso)] = point.close
        except ValueError:
            continue
    return lookup


def fetch_histories_for_lots(core, lots: list[StockCurveLot]) -> dict[tuple[str, str], dict[date, float]]:
    ranges: dict[tuple[str, str], tuple[date, date, str]] = {}
    for lot in lots:
        key = (lot.ticker, lot.currency)
        current = ranges.get(key)
        if current is None:
            ranges[key] = (lot.open_date, lot.end_date, lot.currency_raw)
            continue
        ranges[key] = (min(current[0], lot.open_date), max(current[1], lot.end_date), current[2])
    if not ranges:
        return {}

    cache = load_history_cache()
    securities = cache.get("securities")
    if not isinstance(securities, dict):
        securities = {}
        cache["securities"] = securities

    histories: dict[tuple[str, str], dict[date, float]] = {}
    pending: dict[tuple[str, str], dict[str, object]] = {}
    fetch_jobs: list[tuple[tuple[str, str], str, str, date, date]] = []
    for key, (start, end, currency_raw) in ranges.items():
        ticker, currency = key
        cache_key = history_cache_key(ticker, currency_raw)
        entry = securities.get(cache_key)
        cached_points = entry_points(entry)
        if entry_covers(cached_points, start, end, entry):
            histories[key] = history_lookup(slice_history_points(cached_points, start, end))
            continue

        fetch_ranges: list[tuple[date, date]] = []
        if not cached_points:
            fetch_ranges.append((start, end))
        else:
            first_cached = date.fromisoformat(cached_points[0].iso)
            last_cached = date.fromisoformat(cached_points[-1].iso)
            if start < first_cached:
                fetch_ranges.append((start, first_cached - timedelta(days=1)))
            if end > last_cached:
                fetch_ranges.append((last_cached + timedelta(days=1), end))
        pending[key] = {
            "ticker": ticker,
            "currency": currency,
            "currency_raw": currency_raw,
            "start": start,
            "end": end,
            "cache_key": cache_key,
            "cached_points": cached_points,
            "fetched_points": [],
        }
        for range_start, range_end in fetch_ranges:
            if range_end >= range_start:
                fetch_jobs.append((key, ticker, currency_raw, range_start, range_end))

    if fetch_jobs:
        emit_progress("拉取收益曲线", f"准备读取 {len(fetch_jobs)} 段腾讯历史日线，首次会慢一点。", 66)
        completed = 0
        workers = min(HISTORY_WORKERS, len(fetch_jobs))
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(fetch_security_history_points_online, core, ticker, currency_raw, start, end): (key, ticker)
                for key, ticker, currency_raw, start, end in fetch_jobs
            }
            for future in as_completed(future_map):
                key, ticker = future_map[future]
                try:
                    points = future.result()
                except Exception:
                    points = []
                pending_item = pending.get(key)
                if pending_item is not None:
                    fetched_points = pending_item.setdefault("fetched_points", [])
                    if isinstance(fetched_points, list):
                        fetched_points.extend(points)
                completed += 1
                emit_progress("拉取收益曲线", f"历史日线 {completed}/{len(fetch_jobs)}：{ticker}", 66 + (completed / len(fetch_jobs)) * 8)

    should_save = bool(pending)
    for key, item in pending.items():
        ticker = str(item["ticker"])
        currency = str(item["currency"])
        currency_raw = str(item["currency_raw"])
        cache_key = str(item["cache_key"])
        start = item["start"]
        end = item["end"]
        cached_points = list(item.get("cached_points") or [])
        fetched_points = list(item.get("fetched_points") or [])
        merged = merge_cache_points(cached_points, fetched_points)
        securities[cache_key] = {
            "ticker": clean_text(ticker),
            "currency": clean_text(currency_raw),
            "fetched_at": date.today().isoformat(),
            "points": serialize_history_points(merged),
        }
        histories[key] = history_lookup(slice_history_points(merged, start, end))
    if should_save:
        save_history_cache(cache)
    return histories


def close_for_day(history: dict[date, float], day: date) -> float | None:
    if day in history:
        return history[day]
    candidates = [history_day for history_day in history if history_day <= day]
    if not candidates:
        return None
    return history[max(candidates)]


def lot_price_pnl(lot: StockCurveLot, entry_price: float, price: float, fee: float | None = None) -> float:
    gross_qty = abs(lot.quantity)
    charge = lot.fee if fee is None else fee
    if lot.is_short:
        return (entry_price - price) * gross_qty - charge
    return (price - entry_price) * gross_qty - charge


def unrealized_pnl(lot: StockCurveLot, price: float, entry_price: float | None = None) -> float:
    fee = lot.entry_fee if lot.entry_fee is not None else lot.fee
    return lot_price_pnl(lot, entry_price if entry_price is not None else lot.open_price, price, fee)


def mark_price_for_day(lot: StockCurveLot, day: date, history: dict[date, float]) -> float | None:
    return close_for_day(history, day)


def entry_price_for_curve(lot: StockCurveLot, history: dict[date, float]) -> float:
    return close_for_day(history, lot.open_date) or lot.open_price


def realized_pnl_for_curve(lot: StockCurveLot, history: dict[date, float]) -> float:
    if lot.realized_pnl is not None:
        return float(lot.realized_pnl)
    close_price = lot.close_price
    if close_price is None and lot.close_date is not None:
        close_price = close_for_day(history, lot.close_date)
    if close_price is None:
        return 0.0
    return lot_price_pnl(lot, lot.open_price, close_price)


def current_holding_price_lookup(data: dict[str, object] | None) -> dict[tuple[str, str], float]:
    if not isinstance(data, dict):
        return {}
    prices: dict[tuple[str, str], float] = {}
    for item in data.get("holdings", []) or []:
        if not isinstance(item, dict):
            continue
        ticker = clean_text(item.get("ticker"))
        currency = clean_text(item.get("currency"))
        price = parse_display_number(item.get("last_price"))
        if ticker and currency and price is not None and price > 0:
            prices[(ticker, currency)] = float(price)
    return prices


def mark_price_for_curve_day(
    lot: StockCurveLot,
    day: date,
    history: dict[date, float],
    current_prices: dict[tuple[str, str], float],
) -> float | None:
    if day >= date.today():
        price = current_prices.get((lot.ticker, lot.currency))
        if price is not None:
            return price
    return mark_price_for_day(lot, day, history)


def point_dates_for_currency(lots: list[StockCurveLot], histories: dict[tuple[str, str], dict[date, float]], events: list[RealizedCurveEvent]) -> dict[str, set[date]]:
    by_currency: dict[str, set[date]] = {}
    for lot in lots:
        days = by_currency.setdefault(lot.currency, set())
        history = histories.get((lot.ticker, lot.currency), {})
        for day in history:
            if lot.open_date <= day <= lot.end_date:
                days.add(day)
        days.add(lot.open_date)
        if lot.close_date:
            days.add(lot.close_date)
        elif lot.open_date <= date.today():
            days.add(date.today())
    for event in events:
        by_currency.setdefault(event.currency, set()).add(event.event_date)
    return by_currency


def build_historical_curve_series(core, rows: list[tuple[int, dict[int, object]]], data: dict[str, object] | None = None) -> list[dict[str, object]]:
    lots = build_stock_lots(core, rows)
    if not lots:
        return []
    non_stock_events = build_non_stock_realized_events(core, rows)
    lots = adjust_stock_lot_realized_to_summary(lots, data, non_stock_events)
    residual_events = residual_realized_events_from_summary(core, rows, data, lots, non_stock_events)
    histories = fetch_histories_for_lots(core, lots)
    events = [*non_stock_events, *residual_events, *build_dividend_realized_events(core)]
    current_prices = current_holding_price_lookup(data)
    dates_by_currency = point_dates_for_currency(lots, histories, events)
    series_list: list[dict[str, object]] = []
    for currency, days in sorted(dates_by_currency.items()):
        if not days:
            continue
        currency_lots = [lot for lot in lots if lot.currency == currency]
        currency_events = [event for event in events if event.currency == currency]
        points: list[dict[str, object]] = []
        for day in sorted(days):
            realized_total = sum(
                realized_pnl_for_curve(lot, histories.get((lot.ticker, lot.currency), {}))
                for lot in currency_lots
                if lot.close_date and lot.close_date <= day
            )
            realized_total += sum(event.pnl for event in currency_events if event.event_date <= day)
            unrealized_total = 0.0
            active_capital = 0.0
            has_active_position = False
            has_priced_position = False
            closed_on_day = any(lot.close_date == day for lot in currency_lots)
            event_on_day = any(event.event_date == day for event in currency_events)
            for lot in currency_lots:
                if not (lot.open_date <= day and (lot.close_date is None or day < lot.close_date)):
                    continue
                has_active_position = True
                active_capital += lot.capital
                if day <= lot.open_date:
                    continue
                history = histories.get((lot.ticker, lot.currency), {})
                price = mark_price_for_curve_day(lot, day, history, current_prices)
                if price is None:
                    continue
                has_priced_position = True
                unrealized_total += unrealized_pnl(lot, price)
            if not has_active_position and not has_priced_position and not closed_on_day and not event_on_day:
                continue
            point = {
                "date": date_label(day),
                "iso": day.isoformat(),
                "serial": excel_serial(day),
                "value": realized_total + unrealized_total,
                "float_value": unrealized_total,
                "realized_value": realized_total,
                "total_value": realized_total + unrealized_total,
            }
            if active_capital > EPSILON:
                point["capital"] = active_capital
            elif closed_on_day:
                point["capital"] = 0.0
            points.append(point)
        if len(points) >= 2:
            capital = parse_float(points[-1].get("capital"))
            if capital is None or capital <= EPSILON:
                capital = max((parse_float(point.get("capital")) or 0.0 for point in points), default=0.0)
            if capital <= EPSILON:
                capital = sum(lot.capital for lot in currency_lots)
            series_list.append(
                {
                    "currency": currency,
                    "code": currency,
                    "capital": capital,
                    "points": points,
                    "source": "history",
                }
            )
    return series_list


def replace_curve_series_with_historical_prices(core, rows: list[tuple[int, dict[int, object]]], data: dict[str, object]) -> dict[str, object]:
    started_at = time.monotonic()
    series = build_historical_curve_series(core, rows, data)
    if series:
        data["curve_series"] = series
        total_points = sum(len(item.get("points", []) or []) for item in series if isinstance(item, dict))
        emit_progress("收益曲线", f"真实历史行情曲线完成：{total_points} 个日线点。", 76)
    elif time.monotonic() - started_at > 1:
        emit_progress("收益曲线", "历史行情暂不可用，保留账本原始收益曲线。", 76)
    return data
